import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

server_dir = "/home/reshavsharma/Downloads/code/"  # Path to your server directory
server_stats = []

# Collect data
for server_name in os.listdir(server_dir):
    server_path = os.path.join(server_dir, server_name)
    
    if not os.path.isdir(server_path):
        continue

    stats = {"Server Name": server_name}

    # 1. CPU Load
    cpu_file = os.path.join(server_path, "CPU_Load.csv")
    if os.path.exists(cpu_file):
        df_cpu = pd.read_csv(cpu_file)
        if 'Total(RAW)' in df_cpu.columns:
            stats["CPU Min (%)"] = round(df_cpu['Total(RAW)'].min(), 2)
            stats["CPU Avg (%)"] = round(df_cpu['Total(RAW)'].mean(), 2)
            stats["CPU Max (%)"] = round(df_cpu['Total(RAW)'].max(), 2)
        else:
            print(f"Warning: 'Total(RAW)' not found in {cpu_file}")
    else:
        print(f"Warning: Missing CPU file for {server_name}")

    # 2. Memory Utilization
    memory_file = os.path.join(server_path, "Memory.csv")
    if os.path.exists(memory_file):
        df_mem = pd.read_csv(memory_file)
        if 'Percent Available Memory(RAW)' in df_mem.columns:
            df_mem['Memory Utilized (%)'] = 100 - df_mem['Percent Available Memory(RAW)']
            stats["Memory Min (%)"] = round(df_mem['Memory Utilized (%)'].min(), 2)
            stats["Memory Avg (%)"] = round(df_mem['Memory Utilized (%)'].mean(), 2)
            stats["Memory Max (%)"] = round(df_mem['Memory Utilized (%)'].max(), 2)
        else:
            print(f"Warning: 'Percent Available Memory(RAW)' not found in {memory_file}")
    else:
        print(f"Warning: Missing Memory file for {server_name}")

    # 3. Storage
    storage_file = os.path.join(server_path, "Storage.csv")
    if os.path.exists(storage_file):
        df_storage = pd.read_csv(storage_file)
        if all(col in df_storage.columns for col in ['Free Bytes(RAW)', 'Total(RAW)']):
            last_record = df_storage.iloc[-2]
            total_gb = last_record['Total(RAW)'] / (1024**3)
            free_gb = last_record['Free Bytes(RAW)'] / (1024**3)
            used_gb = total_gb - free_gb
            stats["Storage Total (GB)"] = round(total_gb, 2)
            stats["Storage Used (GB)"] = round(used_gb, 2)
            stats["Storage Avail (GB)"] = round(free_gb, 2)
        else:
            print(f"Warning: Required storage columns not found in {storage_file}")
    else:
        print(f"Warning: Missing Storage file for {server_name}")

    # 4. Uptime
    uptime_file = os.path.join(server_path, "uptime.csv")
    if os.path.exists(uptime_file):
        df_uptime = pd.read_csv(uptime_file)
        if 'System Uptime(RAW)' in df_uptime.columns:
            last_uptime_seconds = df_uptime['System Uptime(RAW)'].iloc[-2]
            uptime_days = last_uptime_seconds / (24 * 3600)
            stats["Uptime (Days)"] = round(uptime_days, 0)
        else:
            print(f"Warning: 'System Uptime(RAW)' not found in {uptime_file}")
    else:
        print(f"Warning: Missing Uptime file for {server_name}")

    server_stats.append(stats)

# Create a DataFrame from the stats
stats_df = pd.DataFrame(server_stats)

# Save to Excel using Pandas ExcelWriter with openpyxl engine
output_file = "Server_full_stats_aesthetic.xlsx"
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    stats_df.to_excel(writer, index=False, sheet_name="Server Stats")

    # Apply formatting
    workbook = writer.book
    sheet = workbook['Server Stats']

    # Style headers (bold, centered, gradient background)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Apply borders around each cell
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in sheet.iter_rows(min_row=2, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            cell.border = border_style
            cell.alignment = Alignment(horizontal="center", vertical="center")  # Center data

    # Add conditional formatting (e.g., highlight high CPU usage)
    for row in sheet.iter_rows(min_row=2, min_col=2, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            if isinstance(cell.value, (int, float)):  # Only format numeric cells
                # Example: highlight CPU usage over 80%
                if "CPU" in sheet.cell(row=1, column=cell.column).value and cell.value > 80:
                    cell.fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")  # Tomato color for high CPU

    # Add soft row shading (alternating row colors for better readability)
    for row_idx in range(2, sheet.max_row + 1):
        if row_idx % 2 == 0:  # Even rows shaded lightly
            for cell in sheet[row_idx]:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Soft gray background

    # Adjust column widths to fit content
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

# Save the Excel file
print(f"Server stats have been saved to {output_file}")

